"""
案分析レポート生成ツール

時間枠一覧のExcelファイルを読み込み、以下の3シートのレポートを自動生成する。
  1. エリア別合計: 全見積指標のエリア別合計
  2. 局別 19-24時構成比: エリア×放送局別の見積ALL（全体 vs 19-24時）
  3. 比別合計: エリア×日にち×タイムランク別の見積ALL

見積カラムはヘッダーから「見積」を含む列を自動検出する。
"""

import os
import sys
import re
from collections import defaultdict
from datetime import datetime

import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter


# ========== 定数 ==========

# エリア表示順（標準）
AREA_ORDER_STANDARD = [
    '関東', '関西', '名古屋', '福岡', '北海道',
    '青森', '岩手', '秋田', '山形', '仙台',
    '福島', '新潟', '長野', '静岡', '富山',
    '石川', '福井', '鳥・島', '広島', '山口',
    '岡・香', '愛媛', '高知', '佐賀', '長崎',
    '熊本', '大分', '宮崎', '鹿児島', '沖縄',
]

RANK_ORDER = ['SB', 'A', 'B', 'C']

# エリアごとの放送局表示順（標準）
STATION_ORDER_BY_AREA = {
    '関東': ['NTV', 'TBS', 'CXT', 'EX', 'TX'],
    '関西': ['YTV', 'MBS', 'KTV', 'ABC', 'TVO'],
    '名古屋': ['CTV', 'CBC', 'THK', 'NBN', 'TVA'],
    '福岡': ['FBS', 'RKB', 'TNC', 'KBC', 'TVQ'],
    '北海道': ['STV', 'HBC', 'UHB', 'HTB', 'TVH'],
    '青森': ['RAB', 'ATV', 'ABA'],
    '岩手': ['IBC', 'TVI', 'MIT', 'IAT'],
    '秋田': ['ABS', 'AKT', 'AAB'],
    '山形': ['YBC', 'YTS', 'TUY', 'SAY'],
    '仙台': ['TBC', 'OXT', 'MMT', 'KHB'],
    '福島': ['FTV', 'FCT', 'KFB', 'TUF'],
    '新潟': ['BSN', 'NST', 'TNY', 'UX'],
    '長野': ['SBC', 'NBS', 'TSB', 'ABN'],
    '静岡': ['SBS', 'SUT', 'SAT', 'SDT'],
    '富山': ['KNB', 'BBT', 'TUT'],
    '石川': ['MRO', 'ITC', 'KTK', 'HAB'],
    '福井': ['FBC', 'FTB'],
    '鳥・島': ['BSS', 'NKT', 'TSK'],
    '広島': ['RCC', 'HTV', 'HOM', 'TSS'],
    '山口': ['KRY', 'TYS', 'YAB'],
    '岡・香': ['RSK', 'OHK', 'TSC', 'RNC', 'KSB'],
    '愛媛': ['RNB', 'EBC', 'ITV', 'EAT'],
    '高知': ['RKC', 'KUT', 'KSS'],
    '佐賀': ['STS'],
    '長崎': ['NBC', 'KTN', 'NCC', 'NIB'],
    '熊本': ['RKK', 'TKU', 'KKT', 'KAB'],
    '大分': ['OBS', 'TOS', 'OAB'],
    '宮崎': ['MRT', 'UMK'],
    '鹿児島': ['MBC', 'KTS', 'KKB', 'KYT'],
    '沖縄': ['RBC', 'OTV', 'QAB'],
}


def sort_stations(area, station_dict):
    """エリアに応じた標準順で局をソートし、未知の局は末尾に追加する。"""
    order = STATION_ORDER_BY_AREA.get(area, [])
    known = [s for s in order if s in station_dict]
    unknown = [s for s in station_dict if s not in order]
    return known + sorted(unknown)

# ========== スタイル ==========

TITLE_FONT = Font(bold=True, size=14)
SUBTITLE_FONT = Font(bold=True, size=12)
HEADER_FILL = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
HEADER_FONT = Font(bold=True, size=11, color='FFFFFF')
SUBTOTAL_FILL = PatternFill(start_color='D9E2F3', end_color='D9E2F3', fill_type='solid')
THIN_BORDER = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin'),
)


def apply_header_style(ws, row, cols):
    for c in cols:
        cell = ws.cell(row=row, column=c)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal='center', wrap_text=True)
        cell.border = THIN_BORDER


def apply_data_style(ws, row, cols, is_total=False):
    for c in cols:
        cell = ws.cell(row=row, column=c)
        cell.border = THIN_BORDER
        if is_total:
            cell.fill = SUBTOTAL_FILL
            cell.font = Font(bold=True)


def set_number_format(ws, row, cols, fmt='#,##0.0'):
    for c in cols:
        ws.cell(row=row, column=c).number_format = fmt


# ========== データ読み込み ==========

def detect_columns(header_row):
    """ヘッダー行からカラムのインデックスを自動検出する。"""
    cols = {
        'area': None,
        'station': None,
        'date': None,
        'start_time': None,
        'rank': None,
    }
    estimate_cols = []  # [(index, display_name), ...]

    for i, val in enumerate(header_row):
        if val is None:
            continue
        # 改行・キャリッジリターン・Excelの _x000D_ を除去
        clean = re.sub(r'(_x000D_|[\r\n\x0d])+', '', str(val)).strip()

        if clean == '地区':
            cols['area'] = i
        elif clean == '放送局':
            cols['station'] = i
        elif clean == '放送日':
            cols['date'] = i
        elif clean == '開始時間':
            cols['start_time'] = i
        elif 'ランク' in clean:
            cols['rank'] = i
        elif '見積' in clean:
            estimate_cols.append((i, clean))

    return cols, estimate_cols


def parse_hour(start_time):
    """開始時間文字列から時間(int)を取得する。"""
    if not start_time:
        return None
    try:
        return int(str(start_time).strip().split(':')[0])
    except (ValueError, IndexError):
        return None


def parse_date(date_val):
    """日付をYYYY/MM/DD文字列に変換する。"""
    if date_val is None:
        return None
    if hasattr(date_val, 'strftime'):
        return date_val.strftime('%Y/%m/%d')
    return str(date_val).strip()


def load_data(filepath):
    """Excelファイルを読み込み、構造化データとメタ情報を返す。"""
    wb = openpyxl.load_workbook(filepath)
    ws = wb[wb.sheetnames[0]]

    # ヘッダー検出
    header = [cell.value for cell in ws[1]]
    cols, estimate_cols = detect_columns(header)

    # 必須カラムチェック
    missing = [k for k in ('area', 'station', 'date', 'start_time', 'rank') if cols[k] is None]
    if missing:
        raise ValueError(f"必須カラムが見つかりません: {missing}")
    if not estimate_cols:
        raise ValueError("見積カラムが見つかりません")

    # データ読み込み
    rows = []
    all_areas = set()

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
        area = row[cols['area']]
        if not area:
            continue
        all_areas.add(area)

        station = row[cols['station']]
        date_str = parse_date(row[cols['date']])
        hour = parse_hour(row[cols['start_time']])
        rank = row[cols['rank']]

        est_vals = {}
        for idx, name in estimate_cols:
            try:
                est_vals[name] = float(row[idx]) if row[idx] is not None else None
            except (ValueError, TypeError):
                est_vals[name] = None

        rows.append({
            'area': area,
            'station': station,
            'date': date_str,
            'hour': hour,
            'rank': rank,
            'estimates': est_vals,
        })

    # エリア順の決定（標準順 + 不足分を末尾に追加）
    area_order = [a for a in AREA_ORDER_STANDARD if a in all_areas]
    for a in sorted(all_areas):
        if a not in area_order:
            area_order.append(a)

    # 見積ALLカラム名を特定（「ALL」を含む見積カラム）
    est_all_name = None
    for _, name in estimate_cols:
        if 'ALL' in name or 'Ａ' in name:
            est_all_name = name
            break
    if est_all_name is None:
        est_all_name = estimate_cols[0][1]  # フォールバック: 最初の見積カラム

    est_names = [name for _, name in estimate_cols]

    return rows, area_order, est_names, est_all_name


# ========== 集計 ==========

def aggregate(rows, area_order, est_names, est_all_name):
    """全集計を実行して結果を返す。"""

    # 1) エリア別 各見積合計
    area_est = defaultdict(lambda: defaultdict(float))
    # 2) エリア×局別 見積ALL (全体 vs 19-24時)
    area_station_total = defaultdict(lambda: defaultdict(float))
    area_station_prime = defaultdict(lambda: defaultdict(float))
    # 3) エリア×日付×ランク別 見積ALL
    area_date_rank = defaultdict(lambda: defaultdict(lambda: defaultdict(float)))

    for r in rows:
        area = r['area']
        station = r['station']
        date_str = r['date']
        hour = r['hour']
        rank = r['rank']
        est = r['estimates']

        # 1) 全見積指標
        for name in est_names:
            v = est.get(name)
            if v is not None:
                area_est[area][name] += v

        # 2) & 3) 見積ALL
        v_all = est.get(est_all_name)
        if v_all is not None:
            area_station_total[area][station] += v_all
            if hour is not None and 19 <= hour < 24:
                area_station_prime[area][station] += v_all
            if date_str and rank:
                area_date_rank[area][date_str][rank] += v_all

    return {
        'area_est': area_est,
        'area_station_total': area_station_total,
        'area_station_prime': area_station_prime,
        'area_date_rank': area_date_rank,
    }


# ========== レポート書き出し ==========

def write_sheet_area_total(wb, area_order, est_names, agg):
    """Sheet1: エリア別 全見積指標合計"""
    ws = wb.active
    ws.title = 'エリア別合計'
    ws.cell(row=1, column=1, value='エリア別 見積指標 合計').font = TITLE_FONT

    # ヘッダー
    ws.cell(row=3, column=1, value='エリア')
    for i, name in enumerate(est_names):
        ws.cell(row=3, column=2 + i, value=name)
    hdr_cols = list(range(1, 2 + len(est_names)))
    apply_header_style(ws, 3, hdr_cols)

    # データ
    r = 4
    grand = defaultdict(float)
    for area in area_order:
        ws.cell(row=r, column=1, value=area)
        for i, name in enumerate(est_names):
            v = agg['area_est'][area].get(name, 0)
            grand[name] += v
            ws.cell(row=r, column=2 + i, value=round(v, 1))
        apply_data_style(ws, r, hdr_cols)
        set_number_format(ws, r, list(range(2, 2 + len(est_names))))
        r += 1

    # 合計行
    ws.cell(row=r, column=1, value='合計')
    for i, name in enumerate(est_names):
        ws.cell(row=r, column=2 + i, value=round(grand[name], 1))
    apply_data_style(ws, r, hdr_cols, is_total=True)
    set_number_format(ws, r, list(range(2, 2 + len(est_names))))

    # 列幅
    ws.column_dimensions['A'].width = 14
    for i in range(len(est_names)):
        ws.column_dimensions[get_column_letter(2 + i)].width = 14


def write_sheet_prime_ratio(wb, area_order, est_all_name, agg):
    """Sheet2: エリア×局別 見積ALL 19-24時構成比"""
    ws = wb.create_sheet('局別 19-24時構成比')
    ws.cell(row=1, column=1,
            value=f'エリア×放送局別 {est_all_name}（全体 vs 19時〜24時）').font = TITLE_FONT

    r = 3
    for area in area_order:
        ws.cell(row=r, column=1, value=f'【{area}】').font = SUBTITLE_FONT
        r += 1

        for c, v in enumerate(['放送局', '全体合計', '19-24時', '構成比'], 1):
            ws.cell(row=r, column=c, value=v)
        apply_header_style(ws, r, [1, 2, 3, 4])
        r += 1

        stations = agg['area_station_total'][area]
        a_total = a_prime = 0
        for station in sort_stations(area, stations):
            total = stations[station]
            prime = agg['area_station_prime'][area].get(station, 0)
            pct = prime / total if total > 0 else 0
            a_total += total
            a_prime += prime

            ws.cell(row=r, column=1, value=station)
            ws.cell(row=r, column=2, value=round(total, 1))
            ws.cell(row=r, column=3, value=round(prime, 1))
            ws.cell(row=r, column=4, value=round(pct, 4))
            ws.cell(row=r, column=4).number_format = '0.0%'
            apply_data_style(ws, r, [1, 2, 3, 4])
            set_number_format(ws, r, [2, 3])
            r += 1

        a_pct = a_prime / a_total if a_total > 0 else 0
        ws.cell(row=r, column=1, value='合計')
        ws.cell(row=r, column=2, value=round(a_total, 1))
        ws.cell(row=r, column=3, value=round(a_prime, 1))
        ws.cell(row=r, column=4, value=round(a_pct, 4))
        ws.cell(row=r, column=4).number_format = '0.0%'
        apply_data_style(ws, r, [1, 2, 3, 4], is_total=True)
        set_number_format(ws, r, [2, 3])
        r += 2

    ws.column_dimensions['A'].width = 14
    ws.column_dimensions['B'].width = 14
    ws.column_dimensions['C'].width = 14
    ws.column_dimensions['D'].width = 12


def write_sheet_rank_by_date(wb, area_order, est_all_name, agg):
    """Sheet3: エリア×日にち×タイムランク別 見積ALL"""
    ws = wb.create_sheet('比別合計')
    ws.cell(row=1, column=1,
            value=f'エリア×日にち×タイムランク別 {est_all_name} 合計').font = TITLE_FONT

    num_cols = 2 + len(RANK_ORDER)
    hdr_cols = list(range(1, num_cols + 1))

    r = 3
    for area in area_order:
        dates = agg['area_date_rank'][area]
        date_order = sorted(dates.keys())

        ws.cell(row=r, column=1, value=f'【{area}】').font = SUBTITLE_FONT
        r += 1

        ws.cell(row=r, column=1, value='日付')
        for i, rk in enumerate(RANK_ORDER):
            ws.cell(row=r, column=2 + i, value=rk)
        ws.cell(row=r, column=2 + len(RANK_ORDER), value='合計')
        apply_header_style(ws, r, hdr_cols)
        r += 1

        grand_rank = defaultdict(float)
        for dt in date_order:
            ws.cell(row=r, column=1, value=dt)
            row_total = 0
            for i, rk in enumerate(RANK_ORDER):
                v = dates[dt].get(rk, 0)
                grand_rank[rk] += v
                row_total += v
                if v > 0:
                    ws.cell(row=r, column=2 + i, value=round(v, 1))
                    set_number_format(ws, r, [2 + i])
                else:
                    ws.cell(row=r, column=2 + i, value='-')
            ws.cell(row=r, column=2 + len(RANK_ORDER), value=round(row_total, 1))
            apply_data_style(ws, r, hdr_cols)
            set_number_format(ws, r, [2 + len(RANK_ORDER)])
            r += 1

        ws.cell(row=r, column=1, value='合計')
        g_all = 0
        for i, rk in enumerate(RANK_ORDER):
            v = grand_rank[rk]
            g_all += v
            if v > 0:
                ws.cell(row=r, column=2 + i, value=round(v, 1))
                set_number_format(ws, r, [2 + i])
            else:
                ws.cell(row=r, column=2 + i, value='-')
        ws.cell(row=r, column=2 + len(RANK_ORDER), value=round(g_all, 1))
        apply_data_style(ws, r, hdr_cols, is_total=True)
        set_number_format(ws, r, [2 + len(RANK_ORDER)])
        r += 2

    ws.column_dimensions['A'].width = 14
    for i in range(len(RANK_ORDER) + 1):
        ws.column_dimensions[get_column_letter(2 + i)].width = 12


def generate_report(input_path, output_path=None):
    """メイン処理: 入力Excelからレポートを生成して保存する。"""
    print(f'読み込み中: {input_path}')
    rows, area_order, est_names, est_all_name = load_data(input_path)
    print(f'  データ件数: {len(rows)}')
    print(f'  エリア数: {len(area_order)}')
    print(f'  検出した見積カラム: {est_names}')
    print(f'  見積ALL判定: {est_all_name}')

    agg = aggregate(rows, area_order, est_names, est_all_name)

    wb = openpyxl.Workbook()
    write_sheet_area_total(wb, area_order, est_names, agg)
    write_sheet_prime_ratio(wb, area_order, est_all_name, agg)
    write_sheet_rank_by_date(wb, area_order, est_all_name, agg)

    if output_path is None:
        base = os.path.splitext(os.path.basename(input_path))[0]
        output_dir = os.path.dirname(input_path)
        output_path = os.path.join(output_dir, f'{base}_分析レポート.xlsx')

    wb.save(output_path)
    print(f'レポート保存完了: {output_path}')
    return output_path


# ========== CLI ==========

if __name__ == '__main__':
    if len(sys.argv) < 2:
        print('使い方: python analyzer.py <入力ファイル.xlsx> [出力ファイル.xlsx]')
        sys.exit(1)

    input_file = sys.argv[1]
    output_file = sys.argv[2] if len(sys.argv) >= 3 else None
    generate_report(input_file, output_file)
